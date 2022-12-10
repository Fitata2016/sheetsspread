import requests
from openpyxl import load_workbook
from flask import Flask, request
from flask_api import status


app = Flask(__name__)

FILE_PATH="users_academlo.xlsx"
SHEET = "users_user"


@app.route("/spreadsheet")
def num_users():
    workbook = load_workbook(FILE_PATH, read_only=True)
    sheet = workbook[SHEET]
    count = -1
    for row in sheet.iter_rows():
        count+=1
        
    return {"Numero de Registros":f"{count}"},200

if __name__== "__main__":
    app.run(host="127.0.0.1", port = 9002, debug= True)